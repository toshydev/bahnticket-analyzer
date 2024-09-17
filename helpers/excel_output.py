import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
from openpyxl.drawing.image import Image

def create_excel(parsed_results, output_file='output.xlsx'):
    """
    Create an Excel file from the parsed results, with formatted columns (date and price),
    a sum at the end of the price column, and appropriate column widths.
    """
    # Convert the parsed results into a DataFrame
    data = []
    for result in parsed_results:
        data.append({
            'Filename': result['filename'],
            'Booking Date': result['parsed_data']['Booking Date'],
            'Date': result['parsed_data']['Date'],
            'Start': result['parsed_data']['Start'],
            'Destination': result['parsed_data']['Destination'],
            'Passenger Count': result['parsed_data']['Passenger Count'],
            'Price': result['parsed_data']['Price'],
            'Ticket Number': result['parsed_data']['Ticket Number']
        })

    df = pd.DataFrame(data)

    # Save the DataFrame to an Excel file
    df.to_excel(output_file, index=False, engine='openpyxl')

    # Load the workbook to modify it (adding sum and formatting)
    wb = load_workbook(output_file)
    ws = wb.active

    # Set column width for readability
    column_widths = {
        'A': 15,  # Filename
        'B': 15,  # Booking Date
        'C': 15,  # Date
        'D': 20,  # Start
        'E': 20,  # Destination
        'F': 18,  # Passenger Count
        'G': 15,  # Price
        'H': 20   # Ticket Number
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define the number formats for date and price columns
    for row in range(2, len(df) + 2):  # Skip header row
        # Set date type for "Booking Date" and "Date" columns (B and C)
        ws[f'B{row}'].number_format = 'DD.MM.YYYY'
        ws[f'C{row}'].number_format = 'DD.MM.YYYY'

        # Set currency type for "Price" column (G) in euros, the values are of type float and should be formatted as currency in the format 9999,99 €
        ws[f'G{row}'].number_format = '#,##0.00 €'

    # Add a sum at the end of the price column
    price_column = 'G'  # Assuming 'Price' is in column G
    total_row = len(df) + 2  # Position for the total sum row

    # Write the sum formula and set the format to currency
    ws[f'{price_column}{total_row}'] = f'=SUM({price_column}2:{price_column}{total_row - 1})'
    ws[f'{price_column}{total_row}'].number_format = '#,##0.00 €'
    
    # Add label for total
    ws[f'F{total_row}'] = 'Total:'
    ws[f'F{total_row}'].alignment = Alignment(horizontal='right')

    # Save the formatted Excel file
    wb.save(output_file)
